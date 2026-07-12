"""Mapping-assisted spreadsheet importer.

Nearly every farmer already runs his operation out of Excel; making him
retype it is the most likely reason he abandons the app in hour one. The
deal (per spec — fully-automatic inference is explicitly out of scope):

  1. He uploads a workbook. We summarize each tab (headers, sample rows,
     and formatting signals — the blue-font-means-input convention is a
     SIGNAL, never a requirement).
  2. The reasoning-tier model PROPOSES a mapping.
  3. He confirms or corrects it once; the mapping is saved by content hash
     so re-import is one tap.
  4. Import runs; anything unresolvable lands as warnings, never silently
     guessed.

v1 targets: crop-plan/rotation tabs → crop_year rows; transaction tabs →
money_transaction; budget tabs → budget_line.
"""
from __future__ import annotations

import hashlib
import io
from datetime import date, datetime, timezone

from sqlalchemy import select
from sqlalchemy.orm import Session

from .. import llm
from ..models import BudgetLine, CropYear, Field, MoneyTransaction, WorkbookMapping

# FSA crop codes for the common Iowa row-crop names
CROP_CODES = {
    "corn": "0041",
    "soybeans": "0081",
    "beans": "0081",
    "soybean": "0081",
    "wheat": "0011",
    "oats": "0016",
    "hay": "0102",
    "alfalfa": "0027",
}

MAPPING_SYSTEM = """You map a farmer's Excel workbook to a farm-records schema.
You see, per tab: the sheet name, the first rows (with cell values), and
which columns use blue fonts (a common farmer convention for input cells —
treat it as a hint, not a rule).

Tab kinds:
- crop_plan: rows are fields; columns include a field name, maybe acres,
  and one column per crop year holding the crop grown ("Corn"/"Beans").
- transactions: dated money rows (date, description, amount, maybe category/crop).
- budget: per-crop cost categories, usually $/acre.
- ignore: anything else (notes, charts, scenario selectors, summaries).

Return ONLY JSON:
{"tabs": [{
  "sheet": "<name>",
  "kind": "crop_plan|transactions|budget|ignore",
  "header_row": <1-based row number of the header>,
  "columns": {"<letter>": "<role>"},
  "crop_year": <year for budget tabs if determinable, else null>,
  "notes": "<one line on anything odd>"
}]}

Column roles by kind:
- crop_plan: field_name, acres, year:<YYYY> (e.g. "year:2026"), ignore
- transactions: date, description, amount, category, kind, crop, ignore
- budget: crop, category, amount_per_acre, ignore
Only include columns you are mapping; omit ignored ones."""


def sha256_bytes(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def summarize_workbook(content: bytes, sample_rows: int = 8) -> list[dict]:
    """Compact per-tab summary for the mapping proposal."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    wb_styles = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    tabs = []
    for name in wb.sheetnames:
        ws = wb[name]
        ws_style = wb_styles[name]
        rows = []
        for r in range(1, min(sample_rows, ws.max_row) + 1):
            row = {}
            for c in range(1, min(ws.max_column, 20) + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    row[get_column_letter(c)] = str(v)[:60]
            rows.append(row)
        blue_cols = set()
        for r in range(1, min(sample_rows * 3, ws_style.max_row) + 1):
            for c in range(1, min(ws_style.max_column, 20) + 1):
                cell = ws_style.cell(row=r, column=c)
                color = getattr(cell.font, "color", None)
                rgb = getattr(color, "rgb", None) if color else None
                if isinstance(rgb, str) and rgb[-6:].upper() in ("0000FF", "0070C0", "1F4E79", "2E75B6"):
                    blue_cols.add(get_column_letter(c))
        tabs.append({"sheet": name, "sample_rows": rows, "blue_font_columns": sorted(blue_cols),
                     "total_rows": ws.max_row})
    return tabs


def propose_mapping(session: Session, content: bytes, cap_usd: float) -> dict:
    import json

    summary = summarize_workbook(content)
    result = llm.complete(
        session,
        purpose="spreadsheet_mapping",
        system=MAPPING_SYSTEM,
        messages=[{"role": "user", "content": json.dumps({"tabs": summary})}],
        max_tokens=2048,
        cap_usd=cap_usd,
    )
    proposal = llm.extract_json(result.text)
    if not isinstance(proposal, dict) or not isinstance(proposal.get("tabs"), list):
        raise ValueError("model returned an unusable mapping proposal")
    return proposal


# ------------------------------------------------------------------ import execution


def _resolve_field(session: Session, name: str) -> Field | None:
    if not name:
        return None
    fields = session.scalars(select(Field).where(Field.archived_at.is_(None))).all()
    lowered = name.strip().lower()
    for f in fields:
        if f.name and f.name.strip().lower() == lowered:
            return f
    for f in fields:  # loose contains match ("Home 80" vs "home eighty" won't match — that's correct)
        if f.name and (lowered in f.name.lower() or f.name.lower() in lowered):
            return f
    return None


def _crop_code(crop_name: str) -> tuple[str, str] | None:
    key = crop_name.strip().lower()
    for name, code in CROP_CODES.items():
        if name in key:
            return code, name if name != "beans" and name != "soybean" else "soybeans"
    return None


def _parse_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _parse_amount(value) -> float | None:
    if isinstance(value, (int, float)):
        return abs(float(value))
    if isinstance(value, str):
        cleaned = value.replace("$", "").replace(",", "").replace("(", "-").replace(")", "").strip()
        try:
            return abs(float(cleaned))
        except ValueError:
            return None
    return None


def run_import(session: Session, wb_row: WorkbookMapping, content: bytes) -> dict:
    import openpyxl

    mapping = wb_row.mapping or {}
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    created = {"crop_years": 0, "transactions": 0, "budget_lines": 0}
    warnings: list[str] = []

    for tab in mapping.get("tabs", []):
        kind = tab.get("kind")
        sheet = tab.get("sheet")
        if kind in (None, "ignore") or sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        header_row = int(tab.get("header_row") or 1)
        columns: dict[str, str] = tab.get("columns") or {}

        for r in range(header_row + 1, ws.max_row + 1):
            values = {role: ws[f"{col}{r}"].value for col, role in columns.items()}
            if all(v in (None, "") for v in values.values()):
                continue
            source = {"workbook_sha": wb_row.content_sha256, "sheet": sheet, "row": r}

            if kind == "crop_plan":
                field = _resolve_field(session, str(values.get("field_name") or ""))
                if field is None:
                    warnings.append(f"{sheet}!r{r}: field '{values.get('field_name')}' not in registry — skipped")
                    continue
                acres = _parse_amount(values.get("acres")) or float(field.clu_calculated_acres or field.gis_acres or 0)
                for role, v in values.items():
                    if not role.startswith("year:") or v in (None, ""):
                        continue
                    year = int(role.split(":")[1])
                    crop = _crop_code(str(v))
                    if crop is None:
                        warnings.append(f"{sheet}!r{r}: unrecognized crop '{v}' for {year} — skipped")
                        continue
                    code, crop_name = crop
                    exists = session.scalar(
                        select(CropYear).where(
                            CropYear.field_id == field.id, CropYear.crop_year == year,
                            CropYear.crop_code == code, CropYear.intended_use_code == "GR",
                        )
                    )
                    if exists is not None:
                        continue
                    session.add(
                        CropYear(field_id=field.id, crop_year=year, crop_code=code,
                                 crop_name=crop_name, reported_acres=acres or 0)
                    )
                    created["crop_years"] += 1

            elif kind == "transactions":
                dup = session.scalar(
                    select(MoneyTransaction).where(MoneyTransaction.source == source)
                )
                if dup is not None:
                    continue
                occurred = _parse_date(values.get("date"))
                amount = _parse_amount(values.get("amount"))
                if occurred is None or amount is None:
                    warnings.append(f"{sheet}!r{r}: missing/unparseable date or amount — skipped")
                    continue
                kind_v = str(values.get("kind") or "").strip().lower()
                session.add(
                    MoneyTransaction(
                        occurred_on=occurred,
                        description=str(values.get("description") or "imported"),
                        kind="income" if kind_v in ("income", "revenue", "sale") else "expense",
                        category=str(values.get("category") or "other"),
                        amount=amount,
                        crop=(str(values["crop"]).lower() if values.get("crop") else None),
                        source=source,
                    )
                )
                created["transactions"] += 1

            elif kind == "budget":
                crop_raw = str(values.get("crop") or "").strip().lower()
                category = str(values.get("category") or "").strip()
                amount = _parse_amount(values.get("amount_per_acre"))
                if not crop_raw or not category or amount is None:
                    warnings.append(f"{sheet}!r{r}: budget row missing crop/category/$ — skipped")
                    continue
                year = int(tab.get("crop_year") or date.today().year)
                exists = session.scalar(
                    select(BudgetLine).where(
                        BudgetLine.crop_year == year, BudgetLine.crop == crop_raw,
                        BudgetLine.category == category,
                    )
                )
                if exists is not None:
                    exists.amount_per_acre = amount
                    exists.source = source
                else:
                    session.add(
                        BudgetLine(crop_year=year, crop=crop_raw, category=category,
                                   amount_per_acre=amount, source=source)
                    )
                    created["budget_lines"] += 1

    wb_row.imported_at = datetime.now(timezone.utc)
    wb_row.import_result = {"created": created, "warnings": warnings[:100]}
    return wb_row.import_result
