"""Spreadsheet importer: synthetic workbook shaped like the Lazy H
reference books (Rotation Plan matrix + Year 1 Budget + a transactions
tab), fake-transport mapping proposal, confirm → import, idempotent
re-import, saved-mapping reuse."""
from __future__ import annotations

import io
import json
import uuid

import pytest

pytestmark = pytest.mark.db


def _build_workbook() -> bytes:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    blue = Font(color="0000FF")

    rot = wb.active
    rot.title = "Rotation Plan"
    rot.append(["Field", "Acres", 2025, 2026])
    rot.append(["WB North 80", 80, "Corn", "Soybeans"])
    rot.append(["WB South 40", 40, "Soybeans", "Corn"])
    rot.append(["Unknown Place", 60, "Corn", "Corn"])  # not in registry → warning
    for row in rot.iter_rows(min_row=2, min_col=3, max_col=4):
        for cell in row:
            cell.font = blue

    bud = wb.create_sheet("Year 1 Budget")
    bud.append(["Crop", "Category", "$/ac"])
    bud.append(["Corn", "Seed", 125.50])
    bud.append(["Corn", "Fertilizer", 210.00])
    bud.append(["Soybeans", "Seed", 62.00])

    txn = wb.create_sheet("Transactions")
    txn.append(["Date", "Description", "Amount", "Category"])
    txn.append(["2026-04-12", "Pioneer seed corn", 12550.00, "Seed"])
    txn.append(["2026-05-02", "Co-op fuel", 1830.25, "Fuel"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


MAPPING = {
    "tabs": [
        {
            "sheet": "Rotation Plan",
            "kind": "crop_plan",
            "header_row": 1,
            "columns": {"A": "field_name", "B": "acres", "C": "year:2025", "D": "year:2026"},
        },
        {
            "sheet": "Year 1 Budget",
            "kind": "budget",
            "header_row": 1,
            "crop_year": 2026,
            "columns": {"A": "crop", "B": "category", "C": "amount_per_acre"},
        },
        {
            "sheet": "Transactions",
            "kind": "transactions",
            "header_row": 1,
            "columns": {"A": "date", "B": "description", "C": "amount", "D": "category"},
        },
    ]
}


def _ensure_fields(app_and_engine):
    from sqlalchemy import select
    from sqlalchemy.orm import Session

    from app.models import Farm, FarmProfile, Field

    _, engine = app_and_engine
    with Session(engine, expire_on_commit=False) as s:
        profile = s.query(FarmProfile).first()
        if profile is None:
            profile = FarmProfile(operation_name="WB Test Farm")
            s.add(profile)
            s.flush()
        farm = s.scalar(select(Farm).where(Farm.farm_number == "8800"))
        if farm is None:
            farm = Farm(farm_profile_id=profile.id, farm_number="8800",
                        state_ansi_code="19", county_ansi_code="153")
            s.add(farm)
            s.flush()
        for i, name in enumerate(["WB North 80", "WB South 40"]):
            exists = s.scalar(select(Field).where(Field.name == name))
            if exists is None:
                s.add(Field(
                    farm_id=farm.id, tract_number="8801", field_number=str(i + 1), name=name,
                    boundary="SRID=4326;MULTIPOLYGON(((-93.8 42.2,-93.79 42.2,-93.79 42.21,-93.8 42.21,-93.8 42.2)))",
                    gis_acres=80 if i == 0 else 40, source="manual",
                ))
        s.commit()


def test_workbook_upload_confirm_import(client, auth_headers, app_and_engine):
    from app import llm

    _ensure_fields(app_and_engine)
    content = _build_workbook()

    llm.set_transport(lambda **kw: (json.dumps(MAPPING), 3000, 600))
    try:
        r = client.post(
            "/api/v1/workbooks",
            files={"file": ("FirstYear_Crop_Plan.xlsx", content,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth_headers,
        )
    finally:
        llm.set_transport(None)
    assert r.status_code == 201, r.text
    wb = r.json()
    assert wb["proposal"]["tabs"][0]["kind"] == "crop_plan"

    r = client.post(f"/api/v1/workbooks/{wb['id']}/confirm", json={"mapping": wb["proposal"]}, headers=auth_headers)
    assert r.status_code == 200, r.text
    result = r.json()["import_result"]
    assert result["created"]["crop_years"] == 4  # 2 fields x 2 years
    assert result["created"]["budget_lines"] == 3
    assert result["created"]["transactions"] == 2
    assert any("Unknown Place" in w for w in result["warnings"])  # honest skip, not a guess

    # crop years landed with FSA codes
    years = client.get("/api/v1/crop-years?year=2026", headers=auth_headers).json()
    wb_years = [c for c in years if c["crop_year"] == 2026 and c["crop_code"] in ("0041", "0081")]
    assert len(wb_years) >= 2

    # re-import same mapping is idempotent
    r = client.post(f"/api/v1/workbooks/{wb['id']}/confirm", json={"mapping": wb["proposal"]}, headers=auth_headers)
    result2 = r.json()["import_result"]
    assert result2["created"] == {"crop_years": 0, "transactions": 0, "budget_lines": 0}

    # same bytes re-uploaded → same row, no new LLM call needed
    r = client.post(
        "/api/v1/workbooks",
        files={"file": ("FirstYear_Crop_Plan.xlsx", content,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth_headers,
    )
    assert r.status_code == 201
    assert r.json()["id"] == wb["id"]


def test_blue_font_signal_is_detected():
    from app.services.workbook_import import summarize_workbook

    tabs = summarize_workbook(_build_workbook())
    rotation = next(t for t in tabs if t["sheet"] == "Rotation Plan")
    assert "C" in rotation["blue_font_columns"]
    assert "D" in rotation["blue_font_columns"]
    assert rotation["sample_rows"][0]["A"] == "Field"
